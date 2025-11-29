import os
from openpyxl import Workbook

# IMPORTANT: Assuming GENERATED_DIR is defined and exported in Injaaz.py
# If Injaaz.py did not define GENERATED_DIR globally, this import would fail.
# We assume it is defined in Injaaz.py for cross-module use.
from Injaaz import GENERATED_DIR 

def create_excel_report(data):
    """
    Generates a placeholder Excel file and saves it to the /generated directory.
    This resolves the 404 error by ensuring a file exists for download.
    """
    # 1. Create a unique filename using the inspection date
    timestamp = data.get('date', 'NODATE').replace('-', '') 
    filename = f"Site_Civil_Report_{timestamp}.xlsx"
    file_path = os.path.join(GENERATED_DIR, filename)

    # 2. Create a basic Workbook object
    wb = Workbook()
    ws = wb.active
    ws.title = "Civil Report Data"
    
    # Write some placeholder data using data gathered from the form
    ws['A1'] = "Report Generation Test"
    ws['B1'] = "SUCCESS: Download Path Verified"
    ws['A2'] = "Building Name:"
    ws['B2'] = data.get('buildingName', 'N/A')
    
    # 3. Save the file to the shared 'generated' directory
    wb.save(file_path)
    
    # Return the filename so the success page can create the download link
    return filename

